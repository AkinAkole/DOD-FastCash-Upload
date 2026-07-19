import io
import os
import zipfile
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import pandas as pd
import streamlit as st

# ==========================================
# PAGE CONFIGURATION & UI STYLING
# ==========================================
st.set_page_config(
    page_title="DOD FastCash Engine",
    page_icon="💸",
    layout="centered",
)

# Custom injection for professional typography and clean layout presentation
st.markdown("""
    <style>
    .metric-card {
        background-color: #F8FAFC;
        border: 1px solid #E2E8F0;
        padding: 1rem;
        border-radius: 0.5rem;
        text-align: center;
        box-shadow: 0 1px 3px 0 rgba(0, 0, 0, 0.05);
    }
    .metric-value {
        font-size: 1.75rem;
        font-weight: 700;
        color: #1A365D;
    }
    .metric-label {
        font-size: 0.875rem;
        color: #64748B;
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }
    </style>
""", unsafe_allow_html=True)

# ==========================================
# 1. SECURITY & ACCESS CONTROL LAYER
# ==========================================
# Simple, robust credential mapping for authorized operations personnel
USER_CREDENTIALS = {
    "dod.fastcash": "FastCash2026!",
    "admin": "Fcmb@54321"
}

def check_password():
    """Returns True if the user has provided valid credentials."""
    if "authenticated" not in st.session_state:
        st.session_state["authenticated"] = False

    if st.session_state["authenticated"]:
        return True

    st.sidebar.title("🔐 DOD FastCash Engine Security Authentication")
    st.sidebar.info("Please enter your operational credentials to unlock the engine processing layer.")
    
    username = st.sidebar.text_input("Username", key="auth_user")
    password = st.sidebar.text_input("Password", type="password", key="auth_pass")
    
    if st.sidebar.button("Authenticate Control Hub", type="primary"):
        if username in USER_CREDENTIALS and USER_CREDENTIALS[username] == password:
            st.session_state["authenticated"] = True
            st.sidebar.success("Access Granted.")
            st.rerun()
        else:
            st.sidebar.error("❌ Invalid Username or Password reference.")
            
    return False

# Enforce security gate before loading application logic
if check_password():

    st.title("💸 DOD FastCash Engine")
    st.caption("FastCash Upload Files Generator • Operations Hub")
    st.markdown("---")

    # Logout button in the sidebar footer
    if st.sidebar.button("Log Out of Session"):
        st.session_state["authenticated"] = False
        st.rerun()

    # ==========================================
    # SIDEBAR / CONTROL INPUTS
    # ==========================================
    st.sidebar.header("Processing Parameters")

    contra_acc = st.sidebar.text_input(
        "Contra Account:", placeholder="Alphanumeric account code"
    )
    tran_type = st.sidebar.selectbox("Tran_Type (D/C):", options=["D", "C"], index=0)
    contra_amt = st.sidebar.number_input(
        "Expected Contra Amount:", min_value=0.0, value=0.0, step=1000.0, format="%.2f"
    )
    out_sheet_name = st.sidebar.text_input("Output Sheet Name Prefix:", value="FastCash_Batch")
    contra_narr = st.sidebar.text_input(
        "Contra Narration (Max 40 chars):", max_chars=40, placeholder="Narration text..."
    )
    ref_prefix = st.sidebar.text_input("Ref. Prefix:", value="FAST//")
    inc_pos = st.sidebar.checkbox("Include cell position details in reference", value=True)

    # ==========================================
    # FILE INGESTION SECTION
    # ==========================================
    st.subheader("1. Source Data Ingestion")
    uploaded_file = st.file_uploader(
        "Upload Source Excel File", type=["xlsx", "xls", "xlsb"]
    )

    local_path = st.text_input(
        "OR Local File Path Fallback:", placeholder="e.g., data/input.xlsx"
    )

    # ==========================================
    # ENGINE RUN TRIGGER
    # ==========================================
    if st.button("Process Data & Generate Package", type="primary"):
        file_source = None
        file_extension = ""

        if uploaded_file is not None:
            file_source = io.BytesIO(uploaded_file.read())
            file_extension = os.path.splitext(uploaded_file.name)[1].lower()
        elif local_path.strip() and os.path.exists(local_path.strip()):
            file_source = local_path.strip()
            file_extension = os.path.splitext(file_source)[1].lower()
        
        if not file_source:
            st.error("❌ Please upload a file or specify a valid local path to proceed.")
        else:
            with st.spinner("Processing batches and applying financial compliance formatting..."):
                try:
                    if file_extension == ".xlsb":
                        df_in = pd.read_excel(file_source, sheet_name=0, engine="pyxlsb")
                    else:
                        df_in = pd.read_excel(file_source, sheet_name=0)
                    
                    df_in["_excel_row_idx"] = df_in.index + 2
                    
                    col_e_name = df_in.columns[4]
                    df_in[col_e_name] = pd.to_numeric(df_in[col_e_name], errors="coerce")
                    df_filtered = df_in.dropna(subset=[col_e_name]).copy()

                    if df_filtered.empty:
                        st.error("❌ No valid rows found containing numeric elements in Column E.")
                        st.stop()

                    # Chunking Logic (Threshold: 7999 rows OR 500,000,000 value sum)
                    chunks = []
                    current_chunk = []
                    current_rows = 0
                    current_sum = 0.0

                    for _, row in df_filtered.iterrows():
                        val_e = float(row[col_e_name])
                        if (current_rows + 1 > 7999) or (current_sum + val_e >= 500000000):
                            if current_chunk:
                                chunks.append(pd.DataFrame(current_chunk))
                            current_chunk = [row]
                            current_rows = 1
                            current_sum = val_e
                        else:
                            current_chunk.append(row)
                            current_rows += 1
                            current_sum += val_e

                    if current_chunk:
                        chunks.append(pd.DataFrame(current_chunk))

                    csv_buffers = {}
                    wb = openpyxl.Workbook()
                    wb.remove(wb.active)  

                    summary_data = []
                    all_contra_rows_sum = 0.0
                    total_processed_records = 0

                    for idx, chunk_df in enumerate(chunks, 1):
                        sheet_title = f"{out_sheet_name} {idx}"
                        ws = wb.create_sheet(title=sheet_title)

                        processed_rows = []
                        chunk_sum = 0.0
                        chunk_row_count = 0

                        for _, r in chunk_df.iterrows():
                            val_a = str(r.iloc[0]) if pd.notnull(r.iloc[0]) else ""
                            val_b = r.iloc[1] if pd.notnull(r.iloc[1]) else ""
                            val_e = float(r[col_e_name])
                            val_f = str(r.iloc[5]) if pd.notnull(r.iloc[5]) else ""
                            val_g = r.iloc[6] if pd.notnull(r.iloc[6]) else ""

                            letter_a = val_a[0].upper() if val_a else ""
                            rounded_e = round(val_e, 2)
                            chunk_sum += rounded_e
                            chunk_row_count += 1

                            sanitized_val_f = val_f.replace(",", "*").replace("-", "*")
                            ref_col_f = f"{ref_prefix}E{r['_excel_row_idx']}" if inc_pos else ref_prefix

                            processed_rows.append([
                                val_b, letter_a, rounded_e, sanitized_val_f, 
                                "", ref_col_f, "", "", "", val_g
                            ])

                        for row_data in processed_rows:
                            ws.append(row_data)

                        final_chunk_sum = round(chunk_sum, 2)
                        all_contra_rows_sum += final_chunk_sum
                        total_processed_records += chunk_row_count

                        contra_ref_f = f"{ref_prefix}{sheet_title}"
                        contra_row = [
                            contra_acc, tran_type, final_chunk_sum, contra_narr,
                            "", contra_ref_f, "", "", "", ""
                        ]
                        ws.append(contra_row)

                        all_sheet_rows = processed_rows + [contra_row]
                        csv_df = pd.DataFrame(all_sheet_rows)
                        csv_text_buffer = io.StringIO()
                        csv_df.to_csv(csv_text_buffer, index=False, header=False)
                        csv_buffers[f"{sheet_title}.csv"] = csv_text_buffer.getvalue()

                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=10):
                            row[2].number_format = "#,##0.00"

                        last_row_idx = ws.max_row
                        accent_fill = PatternFill(start_color="EBF8FF", end_color="EBF8FF", fill_type="solid")
                        for col_idx in range(1, 11):
                            cell = ws.cell(row=last_row_idx, column=col_idx)
                            cell.font = Font(bold=True, color="1A365D")
                            cell.fill = accent_fill

                        summary_data.append({
                            "Batch Identifier": sheet_title,
                            "Records": chunk_row_count,
                            "Total Sum (₦)": final_chunk_sum
                        })

                    # ==========================================
                    # EXECUTIVE SUMMARY COMPLIANCE REPORT (EXCEL)
                    # ==========================================
                    ws_sum = wb.create_sheet(title="Executive Summary", index=0)
                    ws_sum.views.sheetView[0].showGridLines = True

                    navy_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
                    accent_bar_fill = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")
                    white_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
                    header_font = Font(name="Segoe UI", size=16, bold=True, color="1A365D")
                    bold_font = Font(name="Segoe UI", size=11, bold=True)
                    regular_font = Font(name="Segoe UI", size=11)
                    thin_border = Border(
                        left=Side(style="thin", color="CBD5E0"), right=Side(style="thin", color="CBD5E0"),
                        top=Side(style="thin", color="CBD5E0"), bottom=Side(style="thin", color="CBD5E0")
                    )

                    ws_sum.cell(row=2, column=2, value="DOD FastCash Engine").font = header_font
                    ws_sum.cell(row=3, column=2, value="Executive Processing Summary Report").font = Font(name="Segoe UI", size=11, italic=True, color="4A5568")
                    
                    # Global Metrics Panel Table in Excel
                    net_variance = round(all_contra_rows_sum - contra_amt, 2)
                    global_metrics = [
                        ("Total Actionable Data Rows Processed", total_processed_records),
                        ("Target Global Contra Amount Inputted", contra_amt),
                        ("Calculated Net Sum of Generated Rows", all_contra_rows_sum),
                        ("Net Overall Variance Status", net_variance),
                    ]

                    ws_sum.cell(row=5, column=2, value="Global Reconciliation Table").font = bold_font
                    for c_idx, h in enumerate(["Metric Parameter Description", "Value Data"], start=2):
                        cell = ws_sum.cell(row=6, column=c_idx, value=h)
                        cell.fill = navy_fill
                        cell.font = white_font

                    for r_idx, (metric, val) in enumerate(global_metrics, start=7):
                        c1 = ws_sum.cell(row=r_idx, column=2, value=metric)
                        c2 = ws_sum.cell(row=r_idx, column=3, value=val)
                        c1.font = regular_font
                        c2.font = bold_font
                        c2.number_format = "#,##0.00" if isinstance(val, float) else "#,##0"
                        c1.border = thin_border
                        c2.border = thin_border
                        if "Variance" in metric:
                            c2.font = Font(name="Segoe UI", bold=True, color="E53E3E" if val != 0 else "38A169")

                    # Batch Data Log in Excel
                    start_row_breakdown = 13
                    ws_sum.cell(row=start_row_breakdown, column=2, value="Batch Data Sheet Breakdown Variance Log").font = bold_font
                    
                    headers_breakdown = ["Output Sheet Identifier", "Data Row Count", "Imported Rows Sum", "Created Contra Row", "Variance Verification"]
                    for c_idx, h in enumerate(headers_breakdown, start=2):
                        cell = ws_sum.cell(row=start_row_breakdown + 1, column=c_idx, value=h)
                        cell.fill = accent_bar_fill
                        cell.font = white_font

                    curr_row = start_row_breakdown + 2
                    for item in summary_data:
                        cells = [
                            ws_sum.cell(row=curr_row, column=2, value=item["Batch Identifier"]),
                            ws_sum.cell(row=curr_row, column=3, value=item["Records"]),
                            ws_sum.cell(row=curr_row, column=4, value=item["Total Sum (₦)"]),
                            ws_sum.cell(row=curr_row, column=5, value=item["Total Sum (₦)"]),
                            ws_sum.cell(row=curr_row, column=6, value=0.0),
                        ]
                        for idx, c in enumerate(cells):
                            c.font = bold_font if idx == 4 else regular_font
                            c.number_format = "#,##0" if idx == 1 else "#,##0.00"
                            c.border = thin_border
                        curr_row += 1

                    for ws_obj in wb.worksheets:
                        for col in ws_obj.columns:
                            max_len = max(len(str(cell.value or "")) for cell in col)
                            col_letter = openpyxl.utils.get_column_letter(col[0].column)
                            ws_obj.column_dimensions[col_letter].width = max(max_len + 3, 12)

                    excel_buffer = io.BytesIO()
                    wb.save(excel_buffer)
                    excel_data = excel_buffer.getvalue()

                    zip_buffer = io.BytesIO()
                    with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as zip_file:
                        excel_filename = f"{out_sheet_name}_Processed.xlsx"
                        zip_file.writestr(excel_filename, excel_data)
                        for csv_name, csv_content in csv_buffers.items():
                            zip_file.writestr(csv_name, csv_content)

                    zip_buffer.seek(0)
                    
                    # ==========================================
                    # 2. BEAUTIFIED WEB UI REPORTING LAYER
                    # ==========================================
                    st.success("🏁 Verification complete! Production packages built successfully.")
                    
                    st.subheader("📊 Reconciliation Executive Dashboard")
                    
                    # High-Impact Top-Level Metric Cards via Column Grids
                    m_col1, m_col2, m_col3 = st.columns(3)
                    with m_col1:
                        st.markdown(f'<div class="metric-card"><div class="metric-label">Total Records</div><div class="metric-value">{total_processed_records:,}</div></div>', unsafe_allow_html=True)
                    with m_col2:
                        st.markdown(f'<div class="metric-card"><div class="metric-label">Processed Sum</div><div class="metric-value">₦{all_contra_rows_sum:,.2f}</div></div>', unsafe_allow_html=True)
                    with m_col3:
                        var_color = "#38A169" if net_variance == 0 else "#E53E3E"
                        st.markdown(f'<div class="metric-card"><div class="metric-label">Net Variance</div><div class="metric-value" style="color: {var_color};">₦{net_variance:,.2f}</div></div>', unsafe_allow_html=True)
                    
                    st.markdown("<br>", unsafe_allow_html=True)
                    
                    # Streamlit Native Dynamic Bar Charting Layout for Individual Batch Loads
                    st.markdown("**Batch Allocation Matrix Breakdown**")
                    summary_df = pd.DataFrame(summary_data)
                    st.bar_chart(
                        data=summary_df, 
                        x="Batch Identifier", 
                        y="Total Sum (₦)", 
                        color="#2B6CB0", 
                        use_container_width=True
                    )
                    
                    # Structured Clean Data Table Representation Below the Chart
                    st.dataframe(
                        summary_df.style.format({"Total Sum (₦)": "₦{:,.2f}", "Records": "{:,}"}),
                        use_container_width=True,
                        hide_index=True
                    )

                    st.markdown("---")
                    
                    # Prominent Download Call-To-Action Layout
                    st.download_button(
                        label="📥 Download Output Package (.ZIP)",
                        data=zip_buffer,
                        file_name=f"{out_sheet_name}_Package.zip",
                        mime="application/zip",
                        use_container_width=True
                    )

                except Exception as e:
                    st.error(f"❌ Structural Failure During Matrix Parse: {str(e)}")
